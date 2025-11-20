import os
import sqlite3
import random
import re
import shutil
import csv
from kivy.config import Config

# --- 1. 窗口配置 ---
Config.set('graphics', 'width', '400')
Config.set('graphics', 'height', '800')
Config.set('graphics', 'resizable', '0')

# --- 2. 导入模块 ---
from kivy.lang import Builder
from kivy.core.window import Window
from kivy.properties import StringProperty, BooleanProperty, NumericProperty, ListProperty
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.label import Label
from kivy.animation import Animation
from kivy.core.text import LabelBase
from kivy.metrics import dp
from kivymd.app import MDApp
from kivymd.uix.card import MDCard
from kivymd.uix.filemanager import MDFileManager
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.gridlayout import MDGridLayout
from kivymd.uix.dialog import MDDialog
from kivymd.uix.button import MDFlatButton, MDFloatingActionButton
from kivymd.uix.selectioncontrol import MDCheckbox

# --- 3. 字体设置 ---
FONT_PATH = 'font.ttf'
SYSTEM_FONT = 'C:/Windows/Fonts/msyh.ttc'

if not os.path.exists(FONT_PATH):
    if os.path.exists(SYSTEM_FONT):
        try:
            shutil.copyfile(SYSTEM_FONT, FONT_PATH)
        except: pass

REGISTER_PATH = FONT_PATH if os.path.exists(FONT_PATH) else 'Roboto'

LabelBase.register(name='GlobalFont', 
                   fn_regular=REGISTER_PATH, 
                   fn_bold=REGISTER_PATH,
                   fn_italic=REGISTER_PATH,
                   fn_bolditalic=REGISTER_PATH)

# --- 4. 自定义 Toast ---
def show_toast(text):
    label = Label(text=text, font_name='GlobalFont', font_size='16sp', color=(1, 1, 1, 1), padding=(dp(20), dp(10)))
    from kivy.graphics import Color, RoundedRectangle
    def update_rect(instance, value):
        instance.canvas.before.clear()
        with instance.canvas.before:
            Color(0.2, 0.2, 0.2, 0.9)
            RoundedRectangle(pos=instance.pos, size=instance.size, radius=[20,])
    label.bind(pos=update_rect, size=update_rect)
    label.size_hint = (None, None)
    label.size = (dp(200), dp(50))
    label.texture_update()
    label.width = label.texture_size[0] + dp(40)
    label.pos = (Window.width/2 - label.width/2, dp(100))
    label.opacity = 0
    Window.add_widget(label)
    anim = Animation(opacity=1, d=0.2) + Animation(duration=1.0) + Animation(opacity=0, d=0.2)
    anim.bind(on_complete=lambda *x: Window.remove_widget(label))
    anim.start(label)

# --- 5. 数据库管理 ---
class DatabaseManager:
    def __init__(self, db_name='vocab.db'):
        self.conn = sqlite3.connect(db_name)
        self.cursor = self.conn.cursor()
        self.init_db()

    def init_db(self):
        self.cursor.execute("CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT)")
        self.cursor.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('tutorial_seen', '0')")
        self.cursor.execute("CREATE TABLE IF NOT EXISTS libraries (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT, is_active INTEGER DEFAULT 1)")
        self.cursor.execute("CREATE TABLE IF NOT EXISTS words (id INTEGER PRIMARY KEY AUTOINCREMENT, english TEXT, chinese TEXT, status INTEGER DEFAULT 0, library_id INTEGER DEFAULT 1)")
        columns = [i[1] for i in self.cursor.execute("PRAGMA table_info(words)")]
        if 'library_id' not in columns:
            self.cursor.execute("ALTER TABLE words ADD COLUMN library_id INTEGER DEFAULT 1")
            self.cursor.execute("INSERT OR IGNORE INTO libraries (id, name, is_active) VALUES (1, '默认词库', 1)")
        self.conn.commit()

    def add_library(self, name):
        self.cursor.execute("INSERT INTO libraries (name, is_active) VALUES (?, 1)", (name,))
        self.conn.commit()
        return self.cursor.lastrowid

    def get_libraries(self):
        self.cursor.execute("SELECT l.id, l.name, l.is_active, COUNT(w.id) FROM libraries l LEFT JOIN words w ON l.id = w.library_id GROUP BY l.id")
        return self.cursor.fetchall()

    def toggle_library_status(self, lib_id, is_active):
        self.cursor.execute("UPDATE libraries SET is_active = ? WHERE id = ?", (1 if is_active else 0, lib_id))
        self.conn.commit()

    def delete_library(self, lib_id):
        self.cursor.execute("DELETE FROM words WHERE library_id = ?", (lib_id,))
        self.cursor.execute("DELETE FROM libraries WHERE id = ?", (lib_id,))
        self.conn.commit()

    def add_word(self, english, chinese, library_id):
        self.cursor.execute("SELECT id FROM words WHERE english=? AND library_id=?", (english, library_id))
        if not self.cursor.fetchone():
            self.cursor.execute("INSERT INTO words (english, chinese, library_id) VALUES (?, ?, ?)", (english, chinese, library_id))
            self.conn.commit()

    def get_words(self, mode='random', filter_status=[0, 1], limit=None):
        placeholders = ','.join('?' for _ in filter_status)
        query = f"SELECT w.id, w.english, w.chinese, w.status FROM words w JOIN libraries l ON w.library_id = l.id WHERE w.status IN ({placeholders}) AND l.is_active = 1"
        if mode == 'random': query += " ORDER BY RANDOM()"
        else: query += " ORDER BY w.id"
        if limit: query += f" LIMIT {limit}"
        self.cursor.execute(query, filter_status)
        data = self.cursor.fetchall()
        return [{'id': r[0], 'en': r[1], 'cn': r[2], 'status': r[3]} for r in data]

    def get_all_words_by_status(self, status):
        query = "SELECT w.id, w.english, w.chinese, w.status FROM words w JOIN libraries l ON w.library_id = l.id WHERE w.status = ? AND l.is_active = 1 ORDER BY w.id DESC"
        self.cursor.execute(query, (status,))
        data = self.cursor.fetchall()
        return [{'id': r[0], 'en': r[1], 'cn': r[2], 'status': r[3]} for r in data]

    def update_status(self, word_id, status):
        self.cursor.execute("UPDATE words SET status = ? WHERE id = ?", (status, word_id))
        self.conn.commit()
    
    def reset_progress(self):
        self.cursor.execute("UPDATE words SET status = 0")
        self.cursor.execute("UPDATE settings SET value = '0' WHERE key = 'tutorial_seen'")
        self.conn.commit()

    def get_stats(self):
        query = "SELECT w.status, COUNT(*) FROM words w JOIN libraries l ON w.library_id = l.id WHERE l.is_active = 1 GROUP BY w.status"
        self.cursor.execute(query)
        data = dict(self.cursor.fetchall())
        return {'new': data.get(0, 0), 'review': data.get(1, 0), 'mastered': data.get(2, 0)}

    def get_total_count(self):
        query = "SELECT COUNT(*) FROM words w JOIN libraries l ON w.library_id = l.id WHERE l.is_active = 1"
        self.cursor.execute(query)
        res = self.cursor.fetchone()
        return res[0] if res else 0

    def get_setting(self, key):
        self.cursor.execute("SELECT value FROM settings WHERE key = ?", (key,))
        res = self.cursor.fetchone()
        return res[0] if res else None

    def set_setting(self, key, value):
        self.cursor.execute("REPLACE INTO settings (key, value) VALUES (?, ?)", (key, value))
        self.conn.commit()

db = DatabaseManager()

# --- 6. KV 界面设计 ---
KV = '''
#:import FadeTransition kivy.uix.screenmanager.FadeTransition
ScreenManager:
    transition: FadeTransition()
    HomeScreen:
    LibraryScreen:
    ImportScreen:
    StudyScreen:
    DetailScreen:

<TutorialDialogContent>:
    orientation: "vertical"
    spacing: dp(10)
    size_hint_y: None
    height: dp(180)
    MDLabel:
        text: "👈 左滑：标记为【熟知】(不再出现)"
        theme_text_color: "Primary"
    MDLabel:
        text: "👉 右滑：标记为【陌生】(加入复习)"
        theme_text_color: "Primary"
    MDLabel:
        text: "👆 点击：显示/隐藏 单词释义"
        theme_text_color: "Primary"
    Widget:
    MDBoxLayout:
        adaptive_height: True
        spacing: dp(10)
        MDCheckbox:
            id: checkbox
            size_hint: None, None
            size: dp(48), dp(48)
            pos_hint: {'center_y': .5}
        MDLabel:
            text: "不再提示"
            theme_text_color: "Secondary"
            font_style: "Caption"
            pos_hint: {'center_y': .5}

<HomeGridCard@MDCard>:
    orientation: "vertical"
    padding: dp(15)
    spacing: dp(5)
    radius: [15]
    elevation: 3
    ripple_behavior: True
    md_bg_color: (1, 1, 1, 1) if app.theme_cls.theme_style == 'Light' else (0.2, 0.2, 0.2, 1)
    size_hint_y: None
    height: dp(130) 

<HomeScreen>:
    name: 'home'
    on_enter: root.update_stats()
    MDBoxLayout:
        orientation: 'vertical'
        md_bg_color: app.theme_cls.bg_light
        MDBoxLayout:
            orientation: 'vertical'
            size_hint_y: None
            height: dp(180)
            padding: dp(25)
            md_bg_color: app.theme_cls.primary_color
            radius: [0, 0, 30, 30]
            elevation: 4
            MDBoxLayout:
                orientation: 'horizontal'
                size_hint_y: None
                height: dp(40)
                MDLabel:
                    text: "咩哒单词"
                    font_style: "H4"
                    bold: True
                    theme_text_color: "Custom"
                    text_color: 1, 1, 1, 1
                MDBoxLayout:
                    adaptive_width: True
                    spacing: dp(10)
                    MDIconButton:
                        icon: "restart"
                        theme_text_color: "Custom"
                        text_color: 1, 1, 1, 1
                        on_release: root.show_reset_dialog()
                    MDIconButton:
                        icon: "theme-light-dark"
                        theme_text_color: "Custom"
                        text_color: 1, 1, 1, 1
                        on_release: app.switch_theme()
            MDLabel:
                id: progress_text
                text: "加载中..."
                font_style: "H6"
                theme_text_color: "Custom"
                text_color: 1, 1, 1, 0.8
            Widget: 

        MDGridLayout:
            cols: 2
            padding: [dp(20), dp(20), dp(20), dp(10)]
            spacing: dp(15)
            adaptive_height: True
            size_hint_y: None
            HomeGridCard:
                on_release: app.toggle_view_mode()
                md_bg_color: (0.95, 0.95, 1, 1) if app.theme_cls.theme_style == 'Light' else (0.2, 0.2, 0.2, 1)
                MDIcon:
                    icon: "swap-horizontal-bold"
                    theme_text_color: "Primary"
                    font_size: "36sp"
                    halign: "center"
                MDLabel:
                    text: "当前模式"
                    halign: "center"
                    font_style: "Caption"
                    theme_text_color: "Secondary"
                MDLabel:
                    text: "英 - 中" if app.view_mode == 'en_to_cn' else "中 - 英"
                    halign: "center"
                    font_style: "H6"
                    bold: True
                    theme_text_color: "Primary"
            HomeGridCard:
                md_bg_color: app.theme_cls.primary_light if app.theme_cls.theme_style == 'Light' else (0.2, 0.2, 0.2, 1)
                on_release: root.check_and_start_study()
                MDIcon:
                    icon: "play-circle"
                    theme_text_color: "Custom"
                    text_color: app.theme_cls.primary_color if app.theme_cls.theme_style == 'Light' else (0.5, 0.5, 1, 1)
                    font_size: "36sp"
                    halign: "center"
                MDLabel:
                    text: "开始背诵"
                    halign: "center"
                    font_style: "H6"
                    bold: True
                    theme_text_color: "Primary"
                MDLabel:
                    text: "进入学习"
                    halign: "center"
                    font_style: "Caption"
                    theme_text_color: "Secondary"
            HomeGridCard:
                on_release: root.open_detail_view('review')
                MDIcon:
                    icon: "alert-circle-outline"
                    theme_text_color: "Custom"
                    text_color: 1, 0.6, 0, 1
                    font_size: "36sp"
                    halign: "center"
                MDLabel:
                    text: "待复习 (点击查看)"
                    halign: "center"
                    font_style: "Caption"
                    theme_text_color: "Secondary"
                MDLabel:
                    id: stat_review
                    text: "0"
                    halign: "center"
                    font_style: "H5"
                    bold: True
                    theme_text_color: "Custom"
                    text_color: 1, 0.6, 0, 1
            HomeGridCard:
                on_release: root.open_detail_view('mastered')
                MDIcon:
                    icon: "checkbox-marked-circle-outline"
                    theme_text_color: "Custom"
                    text_color: 0, 0.7, 0, 1
                    font_size: "36sp"
                    halign: "center"
                MDLabel:
                    text: "已掌握 (点击查看)"
                    halign: "center"
                    font_style: "Caption"
                    theme_text_color: "Secondary"
                MDLabel:
                    id: stat_mastered
                    text: "0"
                    halign: "center"
                    font_style: "H5"
                    bold: True
                    theme_text_color: "Custom"
                    text_color: 0, 0.7, 0, 1

        MDBoxLayout:
            orientation: "horizontal"
            size_hint_y: None
            height: dp(60)
            padding: [dp(30), dp(0)]
            spacing: dp(20)
            MDLabel:
                text: "本次背诵数量:"
                halign: "right"
                valign: "center"
                size_hint_x: 0.4
                font_style: "Subtitle1"
                theme_text_color: "Primary"
            MDTextField:
                id: batch_input
                text: "20"
                input_filter: "int"
                hint_text: "num"
                mode: "rectangle"
                size_hint_x: 0.3
                pos_hint: {"center_y": .5}
                halign: "center"
                line_color_normal: app.theme_cls.primary_color
        Widget: 

        MDCard:
            size_hint_y: None
            height: dp(80)
            elevation: 0
            radius: [25, 25, 0, 0]
            md_bg_color: (0.9, 0.92, 0.95, 1) if app.theme_cls.theme_style == 'Light' else (0.2, 0.2, 0.2, 1)
            on_release: root.manager.current = 'library'
            ripple_behavior: True
            MDRelativeLayout:
                MDIcon:
                    icon: "bookshelf"
                    font_size: "70sp"
                    theme_text_color: "Custom"
                    text_color: app.theme_cls.primary_color
                    opacity: 0.1
                    pos_hint: {"center_x": .5, "center_y": .4}
                MDLabel:
                    text: "词库管理"
                    halign: "center"
                    font_style: "H6"
                    bold: True
                    theme_text_color: "Primary"
                    pos_hint: {"center_x": .5, "center_y": .5}

<LibraryScreen>:
    name: 'library'
    on_enter: root.load_libraries()
    MDBoxLayout:
        orientation: 'vertical'
        md_bg_color: app.theme_cls.bg_light
        MDBoxLayout:
            size_hint_y: None
            height: dp(60)
            padding: [dp(15), 0]
            md_bg_color: app.theme_cls.primary_color
            elevation: 4
            MDIconButton:
                icon: "arrow-left"
                theme_text_color: "Custom"
                text_color: 1, 1, 1, 1
                on_release: root.manager.current = 'home'
                pos_hint: {"center_y": .5}
            MDLabel:
                text: "我的词库"
                font_style: "H6"
                bold: True
                theme_text_color: "Custom"
                text_color: 1, 1, 1, 1
                pos_hint: {"center_y": .5}
        MDScrollView:
            MDBoxLayout:
                id: lib_container
                orientation: 'vertical'
                adaptive_height: True
                padding: dp(15)
                spacing: dp(15)
    MDFloatingActionButton:
        icon: "plus"
        pos_hint: {"right": .95, "bottom": .05}
        elevation: 4
        on_release: root.manager.current = 'import'

<LibraryItem>:
    orientation: "horizontal"
    size_hint_y: None
    height: dp(80)
    radius: [10]
    elevation: 2
    padding: dp(15)
    spacing: dp(10)
    md_bg_color: (1, 1, 1, 1) if app.theme_cls.theme_style == 'Light' else (0.2, 0.2, 0.2, 1)
    MDIcon:
        icon: "book-open-page-variant"
        font_size: "30sp"
        theme_text_color: "Primary"
        pos_hint: {"center_y": .5}
    MDBoxLayout:
        orientation: "vertical"
        pos_hint: {"center_y": .5}
        MDLabel:
            text: root.lib_name
            font_style: "Subtitle1"
            bold: True
            theme_text_color: "Primary"
        MDLabel:
            text: root.word_count + " 词"
            font_style: "Caption"
            theme_text_color: "Secondary"
    MDSwitch:
        active: root.is_active
        pos_hint: {"center_y": .5}
        on_active: root.toggle_active(self.active)
    MDIconButton:
        icon: "delete"
        theme_text_color: "Error"
        pos_hint: {"center_y": .5}
        on_release: root.delete_library()

<StudyScreen>:
    name: 'study'
    on_enter: root.load_words()
    MDBoxLayout:
        orientation: 'vertical'
        md_bg_color: app.theme_cls.bg_light
        MDBoxLayout:
            size_hint_y: None
            height: dp(60)
            padding: [dp(15), 0]
            md_bg_color: app.theme_cls.primary_color
            elevation: 4
            MDIconButton:
                icon: "arrow-left"
                theme_text_color: "Custom"
                text_color: 1, 1, 1, 1
                on_release: root.manager.current = 'home'
                pos_hint: {"center_y": .5}
            MDLabel:
                id: title_label
                text: "背诵中"
                font_style: "H6"
                bold: True
                theme_text_color: "Custom"
                text_color: 1, 1, 1, 1
                pos_hint: {"center_y": .5}
            MDLabel:
                id: count_label
                text: "0 / 0"
                halign: "right"
                font_style: "H6"
                theme_text_color: "Custom"
                text_color: 1, 1, 1, 0.9
                pos_hint: {"center_y": .5}
        MDScrollView:
            MDBoxLayout:
                id: list_container
                orientation: 'vertical'
                adaptive_height: True
                padding: dp(15)
                spacing: dp(15)

<DetailScreen>:
    name: 'detail'
    on_enter: root.load_data()
    MDBoxLayout:
        orientation: 'vertical'
        md_bg_color: app.theme_cls.bg_light
        MDBoxLayout:
            size_hint_y: None
            height: dp(60)
            padding: [dp(15), 0]
            md_bg_color: app.theme_cls.primary_color
            elevation: 4
            MDIconButton:
                icon: "arrow-left"
                theme_text_color: "Custom"
                text_color: 1, 1, 1, 1
                on_release: root.manager.current = 'home'
                pos_hint: {"center_y": .5}
            MDLabel:
                id: header_title
                text: "词汇列表"
                font_style: "H6"
                bold: True
                theme_text_color: "Custom"
                text_color: 1, 1, 1, 1
                pos_hint: {"center_y": .5}
        MDScrollView:
            MDBoxLayout:
                id: detail_container
                orientation: 'vertical'
                adaptive_height: True
                padding: dp(15)
                spacing: dp(10)

<SimpleWordItem>:
    orientation: "vertical"
    size_hint_y: None
    height: dp(80)
    radius: [10]
    elevation: 1
    padding: dp(10)
    md_bg_color: (1, 1, 1, 1) if app.theme_cls.theme_style == 'Light' else (0.2, 0.2, 0.2, 1)
    MDBoxLayout:
        orientation: "horizontal"
        spacing: dp(10)
        MDBoxLayout:
            orientation: "vertical"
            MDLabel:
                text: root.en_text
                font_style: "H6"
                bold: True
                theme_text_color: "Primary"
            MDLabel:
                text: root.cn_text
                font_style: "Subtitle1"
                theme_text_color: "Secondary"
        MDIconButton:
            icon: "refresh"
            on_release: root.reset_word()
            theme_text_color: "Hint"

<WordListItem>:
    orientation: "vertical"
    size_hint_y: None
    height: dp(110)
    radius: [15]
    elevation: 2
    padding: dp(10)
    ripple_behavior: False
    md_bg_color: (1, 1, 1, 1) if app.theme_cls.theme_style == 'Light' else (0.2, 0.2, 0.2, 1)
    canvas.before:
        Color:
            rgba: root.border_color
        Line:
            width: dp(2)
            rounded_rectangle: self.x, self.y, self.width, self.height, dp(15)
    MDBoxLayout:
        orientation: "vertical"
        MDLabel:
            text: root.main_text
            font_style: "H5"
            bold: True
            halign: "center"
            theme_text_color: "Primary"
            valign: "center"
            size_hint_y: 0.6
        MDLabel:
            text: root.sub_text
            font_style: "Subtitle1"
            halign: "center"
            theme_text_color: "Primary" if root.is_revealed else "Secondary"
            opacity: 1 if root.is_revealed else 0
            valign: "center"
            size_hint_y: 0.4
    MDSeparator:
    MDLabel:
        text: "← 熟知  |  模糊 →"
        font_style: "Overline"
        theme_text_color: "Hint"
        halign: "center"
        size_hint_y: None
        height: dp(20)

<ImportScreen>:
    name: 'import'
    MDBoxLayout:
        orientation: 'vertical'
        md_bg_color: app.theme_cls.bg_light
        MDTopAppBar:
            title: "导入词库"
            left_action_items: [["arrow-left", lambda x: root.go_back()]]
            elevation: 0
            md_bg_color: app.theme_cls.primary_color
        MDBoxLayout:
            orientation: 'vertical'
            padding: dp(30)
            spacing: dp(30)
            Widget: 
            MDIcon:
                icon: "file-upload"
                halign: "center"
                font_size: "100sp"
                theme_text_color: "Custom"
                text_color: app.theme_cls.primary_color
            MDLabel:
                text: "选择 .xlsx 或 .csv 文件"
                halign: "center"
                font_style: "H6"
                theme_text_color: "Primary"
            MDFillRoundFlatIconButton:
                icon: "folder-open"
                text: "浏览文件"
                pos_hint: {"center_x": .5}
                size_hint_x: 0.7
                padding: dp(15)
                on_release: root.file_manager_open()
            MDLabel:
                id: selected_path
                text: "尚未选择文件"
                halign: "center"
                font_style: "Caption"
                theme_text_color: "Secondary"
            MDRaisedButton:
                id: btn_import
                text: "确认导入"
                pos_hint: {"center_x": .5}
                size_hint_x: 0.7
                elevation: 2
                disabled: True
                on_release: root.process_import()
                md_bg_color: app.theme_cls.accent_color
            Widget: 
'''

# --- 7. 组件逻辑 ---

class TutorialDialogContent(MDBoxLayout):
    pass

class LibraryItem(MDCard):
    lib_id = NumericProperty(0)
    lib_name = StringProperty("")
    is_active = BooleanProperty(True)
    word_count = StringProperty("0")
    def __init__(self, lib_data, parent_screen, **kwargs):
        super().__init__(**kwargs)
        self.lib_id = lib_data[0]
        self.lib_name = str(lib_data[1])
        self.is_active = True if lib_data[2] == 1 else False
        self.word_count = str(lib_data[3])
        self.parent_screen = parent_screen
    def toggle_active(self, value):
        db.toggle_library_status(self.lib_id, value)
    def delete_library(self):
        db.delete_library(self.lib_id)
        self.parent_screen.load_libraries()
        show_toast("词库已删除")

class SimpleWordItem(MDCard):
    en_text = StringProperty("")
    cn_text = StringProperty("")
    def __init__(self, word_data, **kwargs):
        super().__init__(**kwargs)
        self.word_data = word_data
        self.en_text = word_data['en']
        self.cn_text = word_data['cn']
    def reset_word(self):
        db.update_status(self.word_data['id'], 0)
        show_toast("已重置为新词")
        if self.parent: self.parent.remove_widget(self)

class WordListItem(MDCard):
    main_text = StringProperty("")
    sub_text = StringProperty("")
    is_revealed = BooleanProperty(False)
    border_color = ListProperty([0, 0, 0, 0])
    def __init__(self, word_data, parent_list, study_screen, **kwargs):
        super().__init__(**kwargs)
        self.word_data = word_data
        self.parent_list = parent_list
        self.study_screen = study_screen
        self.touch_start_pos = (0, 0)
        app = MDApp.get_running_app()
        if app.view_mode == 'en_to_cn':
            self.main_text = word_data['en']
            self.sub_text = word_data['cn']
        else:
            self.main_text = word_data['cn']
            self.sub_text = word_data['en']
        if word_data['status'] == 1: self.border_color = [1, 0.6, 0, 0.5]
    def on_touch_down(self, touch):
        if self.collide_point(*touch.pos): self.touch_start_pos = touch.pos
        return super().on_touch_down(touch)
    def on_touch_up(self, touch):
        if self.collide_point(*touch.pos):
            dx = touch.x - self.touch_start_pos[0]
            dy = touch.y - self.touch_start_pos[1]
            if abs(dx) > abs(dy) and abs(dx) > 60:
                if dx < 0: self.handle_swipe('left')
                else: self.handle_swipe('right')
                return True
            elif abs(dx) < 30 and abs(dy) < 30:
                self.is_revealed = not self.is_revealed
                return True
        return super().on_touch_up(touch)
    def handle_swipe(self, direction):
        anim = Animation(opacity=0, height=0, duration=0.3)
        anim.bind(on_complete=self.remove_self)
        anim.start(self)
        if direction == 'left': db.update_status(self.word_data['id'], 2)
        else: db.update_status(self.word_data['id'], 1)
    def remove_self(self, *args):
        if self.parent:
            self.parent.remove_widget(self)
            self.study_screen.update_progress()

# --- 8. 屏幕逻辑 ---

class HomeScreen(Screen):
    dialog = None
    tutorial_dialog = None
    tutorial_content = None
    def update_stats(self):
        stats = db.get_stats()
        total = db.get_total_count()
        self.ids.stat_mastered.text = str(stats['mastered'])
        self.ids.stat_review.text = str(stats['review'])
        self.ids.progress_text.text = f"已掌握: {stats['mastered']}  /  总词数: {total}"
    def check_and_start_study(self):
        total = db.get_total_count()
        if total == 0:
            show_toast("请先在【词库管理】中添加词库")
            return
        seen = db.get_setting('tutorial_seen')
        if seen == '1': self.start_study_logic()
        else: self.show_tutorial_dialog()
    def show_tutorial_dialog(self):
        if not self.tutorial_dialog:
            self.tutorial_content = TutorialDialogContent()
            app = MDApp.get_running_app()
            self.tutorial_dialog = MDDialog(
                title="操作指引",
                type="custom",
                content_cls=self.tutorial_content,
                buttons=[MDFlatButton(text="开始学习", text_color=app.theme_cls.primary_color, on_release=self.on_tutorial_confirm)],
            )
        self.tutorial_dialog.open()
    def on_tutorial_confirm(self, *args):
        if self.tutorial_content.ids.checkbox.active:
            db.set_setting('tutorial_seen', '1')
        self.tutorial_dialog.dismiss()
        self.start_study_logic()
    def start_study_logic(self):
        try:
            count = int(self.ids.batch_input.text)
            if count <= 0: count = 20
        except: count = 20
        app = MDApp.get_running_app()
        app.batch_limit = count
        self.manager.current = 'study'
    def open_detail_view(self, type_str):
        app = MDApp.get_running_app()
        app.detail_view_type = type_str
        self.manager.current = 'detail'
    def show_reset_dialog(self):
        if not self.dialog:
            self.dialog = MDDialog(
                title="⚠️ 确认重置?",
                text="这会清除所有记忆进度，并重置新手引导。\n此操作不可撤销。",
                buttons=[
                    MDFlatButton(text="取消", on_release=lambda x: self.dialog.dismiss()),
                    MDFlatButton(text="确定重置", text_color=(1, 0, 0, 1), on_release=self.execute_reset)
                ],
            )
        self.dialog.open()
    def execute_reset(self, *args):
        db.reset_progress()
        self.dialog.dismiss()
        self.update_stats()
        show_toast("进度已全部重置！")

class LibraryScreen(Screen):
    def load_libraries(self):
        container = self.ids.lib_container
        container.clear_widgets()
        libs = db.get_libraries()
        if not libs:
            show_toast("暂无词库，请点击右下角添加")
            return
        for lib in libs:
            item = LibraryItem(lib_data=lib, parent_screen=self)
            container.add_widget(item)

class ImportScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.file_manager = MDFileManager(
            exit_manager=self.exit_manager,
            select_path=self.select_path,
            preview=False,
        )
        self.current_path = ""
    def file_manager_open(self):
        self.file_manager.show(os.path.expanduser("~"))
    def select_path(self, path):
        self.exit_manager()
        if path.endswith(('.xlsx', '.xls', '.csv')):
            self.current_path = path
            self.ids.selected_path.text = os.path.basename(path)
            self.ids.btn_import.disabled = False
        else: show_toast("请选择 Excel 或 CSV 文件")
    def exit_manager(self, *args):
        self.file_manager.close()
    def process_import(self):
        try:
            # 使用纯Python库读取，移除Pandas依赖
            from openpyxl import load_workbook
            import xlrd

            rows_data = []
            path = self.current_path
            filename = os.path.basename(path)
            lib_name = os.path.splitext(filename)[0]
            lib_id = db.add_library(lib_name)

            if path.endswith('.csv'):
                try:
                    with open(path, 'r', encoding='utf-8-sig') as f:
                        reader = csv.reader(f)
                        rows_data = list(reader)
                except UnicodeDecodeError:
                    with open(path, 'r', encoding='gbk') as f:
                        reader = csv.reader(f)
                        rows_data = list(reader)
            elif path.endswith('.xlsx'):
                wb = load_workbook(path, data_only=True)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    rows_data.append(row)
            elif path.endswith('.xls'):
                book = xlrd.open_workbook(path)
                sheet = book.sheet_by_index(0)
                for rx in range(sheet.nrows):
                    rows_data.append(sheet.row_values(rx))

            count = 0
            for row in rows_data:
                vals = [str(v).strip() for v in row if v is not None and str(v).strip() != ""]
                if len(vals) < 2: continue
                cn, en = "", ""
                for v in vals:
                    if re.search(r'[\u4e00-\u9fa5]', v): cn = v
                    else: en = v
                if cn and en:
                    db.add_word(en, cn, lib_id)
                    count += 1

            show_toast(f"成功导入: {lib_name} ({count}词)")
            self.manager.current = 'library'
        except Exception as e:
            show_toast(f"导入失败: {str(e)}")
    def go_back(self):
        self.manager.current = 'library'

class DetailScreen(Screen):
    def load_data(self):
        container = self.ids.detail_container
        container.clear_widgets()
        app = MDApp.get_running_app()
        view_type = app.detail_view_type 
        status_code = 1 if view_type == 'review' else 2
        self.ids.header_title.text = "待复习列表" if view_type == 'review' else "已掌握列表"
        words = db.get_all_words_by_status(status_code)
        if not words:
            show_toast("列表为空")
            return
        for word in words[:200]:
            item = SimpleWordItem(word_data=word)
            container.add_widget(item)
        if len(words) > 200: show_toast(f"仅显示前 200 个 (共 {len(words)} 个)")

class StudyScreen(Screen):
    initial_count = 0
    finished_count = 0
    def load_words(self):
        container = self.ids.list_container
        container.clear_widgets()
        app = MDApp.get_running_app()
        limit = app.batch_limit
        words = db.get_words(mode='random', filter_status=[0, 1], limit=limit)
        if not words:
            show_toast("暂无需要背诵的单词")
            self.manager.current = 'home'
            return
        self.initial_count = len(words)
        self.finished_count = 0
        self.update_label()
        for word in words:
            item = WordListItem(word_data=word, parent_list=container, study_screen=self)
            container.add_widget(item)
    def update_progress(self):
        self.finished_count += 1
        if self.finished_count >= self.initial_count:
            show_toast("本组学习完成！")
            self.manager.current = 'home'
        else: self.update_label()
    def update_label(self):
        current = self.finished_count + 1
        self.ids.count_label.text = f"{current} / {self.initial_count}"

# --- 9. 主程序 ---
class VocabApp(MDApp):
    view_mode = StringProperty('en_to_cn')
    batch_limit = NumericProperty(20)
    detail_view_type = StringProperty('')
    def build(self):
        self.title = "咩哒单词"
        self.theme_cls.primary_palette = "Indigo"
        self.theme_cls.accent_palette = "Pink"
        self.theme_cls.theme_style = "Light"
        for style_name, style_values in self.theme_cls.font_styles.items():
            if style_name != 'Icon': style_values[0] = 'GlobalFont'
        Window.bind(on_request_close=self.on_request_close)
        return Builder.load_string(KV)
    def on_start(self):
        if self.root:
            home = self.root.get_screen('home')
            home.update_stats()
    def on_request_close(self, *args):
        return False
    def switch_theme(self):
        self.theme_cls.theme_style = "Dark" if self.theme_cls.theme_style == "Light" else "Light"
    def toggle_view_mode(self):
        if self.view_mode == 'en_to_cn':
            self.view_mode = 'cn_to_en'
            show_toast("模式切换：看中文 -> 背英文")
        else:
            self.view_mode = 'en_to_cn'
            show_toast("模式切换：看英文 -> 背中文")

if __name__ == '__main__':
    VocabApp().run()